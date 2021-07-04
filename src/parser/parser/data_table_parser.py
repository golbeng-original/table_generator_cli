import os

from openpyxl import Workbook, load_workbook
from parser.schema_table_parser import ExcelSchemaData, ExcelSchemaField

class ExcelData:

    __excel_schema_data:ExcelSchemaData = None

    __data_rows = [] # list<list<object>>
    __data_column_fields = [] # list<str>

    def __init__(self, excel_schema_data:ExcelSchemaData, data_column_fileds, data_rows):
        self.__excel_schema_data = excel_schema_data
        self.__data_column_fields = data_column_fileds
        self.__data_rows = data_rows

    def get_excel_schema_data(self) -> ExcelSchemaData:
        return self.__excel_schema_data

    def get_row_count(self) -> int:
        return self.__data_rows.count

    def get_row(self, index) -> int:
        if index >= self.get_row_count():
            return None

        return self.__data_rows[index]

    def get_rows(self) -> list:
        for row in self.__data_rows:
            yield row

    def get_column_mapping_rows(self):
        
        for row in self.__data_rows:

            mapping_row = []
            for schema_field in self.__excel_schema_data.get_fields():
                index = self.get_column_field_index(schema_field.name)
                if index == -1:
                    continue

                mapping_row.append((schema_field.name, row[index]))

            yield mapping_row

    def get_column_fields(self) -> list:
        return self.__data_column_fields

    def get_column_field_index(self, field_name:str) -> int:
        return self.__data_column_fields.index(field_name)



class ExcelDataParser:
    __workbook:Workbook = None
    __worksheet = None

    def parsing(self, excel_schmea_data:ExcelSchemaData, filepath) -> ExcelData:
        
        if os.path.exists(filepath) == False:
            raise FileNotFoundError('{0} is not found'.format(filepath))

        if self.__is_excel_file(filepath) == False:
            raise Exception('{0} is not excel file'.format(filepath))

        try:
            self.__workbook = load_workbook(filename = filepath)
            if 'DATA' not in self.__workbook.sheetnames:
                raise Exception('not found \'DATA\' sheet')

            self.__worksheet = self.__workbook['DATA']

            # data에 있는 schema Column 조사 하기
            column_fileds = self.__parse_data_columns()

            # None 정규화 하기
            data_column_fiels = list(filter(lambda item : item is not None , column_fileds))

            data_rows = self.__parse_data_rows(column_fileds, excel_schmea_data)

            excel_data = ExcelData(excel_schmea_data, data_column_fiels, data_rows)
            return excel_data

        except:
            raise
        finally:
            if self.__workbook is not None:
                self.__workbook.close()

    def __parse_data_columns(self):
        max_column = self.__worksheet.max_column

        data_columns = []

        for colum_index in range(2, max_column + 1):

            cell = self.__worksheet.cell(2, colum_index)
            if cell is None or len(cell.value) == 0:
                data_columns.append(None)
                continue

            column_field_name = cell.value
            column_field_name = column_field_name.strip()

            if column_field_name.startswith('//') is True:
                data_columns.append(None)
                continue

            data_columns.append(column_field_name)

        return data_columns

    def __parse_data_rows(self, column_order_list, schmea_data:ExcelSchemaData):
        
        start_row = 7
        max_row = self.__worksheet.max_row
        max_column = self.__worksheet.max_column

        # SchemaFiled를 조사 하기 위해.. {index, SchemaField}
        column_index_mapping = {}

        data_rows = []

        for row_index in range(start_row, max_row + 1):

            # 무시 체크
            cell = self.__worksheet.cell(row_index, 1)
            cell_str = str(cell.value if cell is not None else '')
            if cell_str.strip().startswith('//') is True:
                continue

            # priamryKey 해당 하는 cell이 정상인가??


            # row 데이터 추출
            row_bundle = []
            for column_index in range(2, max_column + 1):

                # schemaFiled를 조사 및 캐싱한다.
                if column_index not in column_index_mapping:
                    column_field_name = column_order_list[column_index - 2]
                    column_index_mapping[column_index] = schmea_data.find_schema_field(column_field_name)

                column_schema_field:ExcelSchemaField = column_index_mapping[column_index]
                if column_schema_field is None:
                    continue

                cell = self.__worksheet.cell(row_index, column_index)
                value = cell.value if cell is not None else None
                if value is None:
                    value = str(column_schema_field.get_native_default())
                elif isinstance(value, str) and len(value) == 0:
                    value = str(column_schema_field.get_native_default())
                else:
                    value = str(value)

                #print(value)
                row_bundle.append(value)

            data_rows.append(row_bundle)

        return data_rows

    def __is_excel_file(self, filepath):

        filename:str = os.path.basename(filepath)
        return True if filename.endswith('.xlsx') else False