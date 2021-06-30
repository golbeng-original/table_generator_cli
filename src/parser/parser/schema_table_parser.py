import os
from numpy.core import numerictypes
from openpyxl import Workbook, load_workbook
from enum import Enum
import numpy

class ExcelSchemaField:
    name:str = ''
    type:str = ''
    default:str = ''
    title:str = ''
    comment:str = ''
    primary:bool = False

    def get_nativetype(self):

        if len(self.type) == 0:
            return None
        
        lower_type = self.type.lower()

        if lower_type == 'uint':
            return numpy.uint32

        if lower_type == 'int':
            return numpy.int32

        if lower_type == 'bool':
            return bool
        
        if lower_type == 'float':
            return numpy.float32

        if lower_type == 'string':
            return str

        return Enum

    def get_sqlitetype(self):

        native_type = self.get_nativetype()
        if native_type is numpy.int32 or\
            native_type in numpy.uint32 or\
            native_type is bool or\
            native_type is Enum:
            return 'Integer'

        if native_type is numpy.float32:
            return 'Real'

        if native_type is str:
            return 'Text'
    
        return None

    def get_native_default(self):

        native_type = self.get_nativetype()
        if native_type is numpy.uint32:
            return numpy.uint32(self.default)
        
        if native_type is numpy.int32:
            return numpy.int32(self.default)

        if native_type is numpy.float32:
            return numpy.float32(self.default)

        if native_type is bool:
            return True if self.default.lower() == 'true' else False

        return self.default

    def __str__(self):

        output = '[name : {0}]'.format(self.name);
        output += '[type : {0}]'.format(self.type);
        output += '[default : {0}]'.format(self.default);
        output += '[title : {0}]'.format(self.title);
        output += '[comment : {0}]'.format(self.comment);
        output += '[primary : {0}]'.format(self.primary);

        return output;

class ExcelSchemaData:
    schema_name:str = ''
    table_name:str = ''
    db_name:str = ''
    client_file_name:str = ''

    # list[SchemaFiled]
    __fields:list = []

    def __init__(self, schema_name:str, fields):
        self.__fields = fields
        self.schema_name = schema_name
        self.table_name = 'Tbl' + schema_name.capitalize()
        self.db_name = schema_name.capitalize() + '.db'
        self.client_file_name = schema_name.capitalize() + '.byte'

    def find_schema_field(self, fieldname:str):

        if len(self.__fields) == 0:
            return None

        def _filter_func(item:ExcelSchemaField):
            if item.name.lower() == fieldname.lower():
                return item
            return None
            
        result = filter(_filter_func, self.__fields)
        result = list(result)
        if len(result) == 0:
            return None

        return result[0]

    def isprimary(self, fieldname:str):

        if len(self.__fields) == 0:
            return False

        schemafield:ExcelSchemaField = self.find_schema_field(fieldname)
        if schemafield is None:
            return False
        
        return schemafield.primary

    def get_field_count(self):
        return self.__fields.count

    def get_fields(self):
        for field in self.__fields:
            yield field

    def __str__(self):
        print('schema_name = {0}'.format(self.schema_name))
        print('table_name = {0}'.format(self.table_name))
        print('db_name = {0}'.format(self.db_name))
        print('client_file_name = {0}'.format(self.client_file_name))
        for field in self.__fields:
            print('[{0}] - {1}'.format(self.__fields.index(field), field))


class ExcelSchemaParser:
    __workbook:Workbook = None
    __worksheet = None

    # <string, int> = <schemafieldname, filed가 위치하는 Column Index>
    __column_oridinal:dict = {}

    def __del__(self):
        if self.__workbook is None:
            return

        self.__workbook.close()

    def parsing(self, filepath):

        if os.path.exists(filepath) == False:
            raise FileNotFoundError('')

        if self.__is_schema_file(filepath) == False:
            raise Exception('filepath is not schema file')

        try:
            self.__workbook = load_workbook(filename = filepath)

            if 'SCHEMA' not in self.__workbook .sheetnames:
                raise Exception('not found \'SCHEMA\' sheet')
            
            self.__worksheet = self.__workbook ['SCHEMA']
            
            schema_name:str = os.path.basename(filepath)
            findindex = schema_name.lower().index('.schema.xlsx')
            schema_name = schema_name[0: findindex]

            # schema 첫 행 정보 구성하기
            self.__set_column_ordinal()

            # schema 첫행 구성 요소들이 정상인가??
            if self.__is_all_exist_schema_fields() == False:
                raise Exception('not all exist schema fileds')

            # feild 정보 파싱
            schema_fields = self.__get_schema_fields()

            # SchemaData 구성
            schema_data = ExcelSchemaData(schema_name, schema_fields)
            return schema_data

        except:
            raise

    def __is_schema_file(self, filepath):

        filename:str = os.path.basename(filepath)
        filename = filename.lower()
        if filename.endswith('.xlsx') == False:
            return False

        if filename.endswith('.schema.xlsx') == False:
            return False

        return True


    # field의 특성 가져온다. 첫줄에 정의 되어있음..
    # Schema 정보의 field들을 수집
    def __set_column_ordinal(self):
        self.__column_oridinal = {}

        max_column = self.__worksheet.max_column
        for column_index in range(1, max_column+1):
            cell = self.__worksheet.cell(1, column_index)
            if cell is None:
                continue

            self.__column_oridinal[cell.value.lower()] = column_index

    def __get_schema_fields(self):

        schema_fields = []

        max_row = self.__worksheet.max_row
        for row_index in range(2, max_row + 1):
            
            schema_field = ExcelSchemaField()

            # default 필드를 제외하고 모두 조사
            column_keys = [key for key in self.__column_oridinal.keys() if key != 'default']
            for column_key in column_keys:

                column_index = self.__column_oridinal[column_key]

                cell = self.__worksheet.cell(row_index, column_index)

                # Name, Type에 해당하는 cell은 무조건 값이 있어야 한다.
                if column_key == 'name' or column_key == 'type':
                    if cell is None or len(cell.value) == 0:
                        raise Exception('{0} field is empty'.format(column_key))

                # cell.value에 값이 없으면 기본값을 채워주자...
                default_value = str(cell.value) if cell.value is not None else None
                if default_value is None:
                    
                    attr = getattr(schema_field, column_key)
                    if type(attr) is str:
                        default_value = ''
                    elif type(attr) is bool:
                        default_value = False
            
                setattr(schema_field, column_key, default_value)

            # default 조사
            column_keys = [key for key in self.__column_oridinal.keys() if key == 'default']
            for column_key in column_keys:
                
                column_index = self.__column_oridinal[column_key]
                cell = self.__worksheet.cell(row_index, column_index)

                default_value = str(cell.value) if cell.value is not None else ''
                if len(default_value) != 0:
                    setattr(schema_field, column_key, default_value)
                    continue

                if schema_field.get_nativetype() == numpy.int32:
                    default_value = '0'
                elif schema_field.get_nativetype() == numpy.uint32:
                    default_value = '0'
                elif schema_field.get_nativetype() == numpy.float32:
                    default_value = '0'
                elif schema_field.get_nativetype() == bool:
                    default_value = 'False'
                elif schema_field.get_nativetype() == Enum:
                    default_value = 'None'
                else:
                    default_value = ''

                setattr(schema_field, column_key, default_value)

            schema_fields.append(schema_field)

        return schema_fields

    # schema field 들이 정상적으로 구성 되어 있나??
    def __is_all_exist_schema_fields(self):

        check_schema_field = ExcelSchemaField()

        for name in self.__column_oridinal.keys():
            if hasattr(check_schema_field, name) == False:
                return False

        return True