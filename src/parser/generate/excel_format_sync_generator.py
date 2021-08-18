from core.progress_woker import ProgressWorker
import numpy
from openpyxl.cell.cell import Cell
from openpyxl.descriptors import base
from openpyxl.xml.constants import MIN_COLUMN
import generate
import os

from enum import Enum
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from parser.enum_define_parser import EnumMetaData, EnumMetaInfo
from parser.schema_table_parser import ExcelSchemaData, ExcelSchemaField

from core.path_util import convert_path, mkdir_path
from . import excel_format_util as ef_util


class ExcelFormatSyncGenerator:

    __default_row_count = 400

    __enum_data:EnumMetaData = None
    __schema_data:ExcelSchemaData = None

    def __init__(self, enum_data:EnumMetaData, schema_data:ExcelSchemaData):
        self.__enum_data = enum_data
        self.__schema_data = schema_data

    
    def format_sync_async(self, target_excel_data_path:str):
        
        if target_excel_data_path.endswith('.xlsx') == False:
            raise Exception('{0} is not .xlsx file'.format(target_excel_data_path))

        absoulte_path = convert_path(target_excel_data_path)
        mkdir_path(absoulte_path)

        if os.path.exists(absoulte_path) == False:
            raise Exception('{0} is not found'.format(target_excel_data_path))

        worker = ProgressWorker(self.__format_sync(absoulte_path))
        worker.start()

        return worker

    def format_sync_sync(self, target_excel_data_path:str):
        if target_excel_data_path.endswith('.xlsx') == False:
            raise Exception('{0} is not .xlsx file'.format(target_excel_data_path))

        absoulte_path = convert_path(target_excel_data_path)
        mkdir_path(absoulte_path)

        if os.path.exists(absoulte_path) == False:
            raise Exception('{0} is not found'.format(target_excel_data_path))

        return self.__format_sync(absoulte_path)(None)

    def __format_sync(self, target_excel_data_path:str):
        
        def work(worker:ProgressWorker):

            filename = os.path.basename(target_excel_data_path)

            try:
                if worker: worker.updateProgress(0, "{0} loading...".format(filename))

                workbook:Workbook = load_workbook(target_excel_data_path)

                if worker: worker.updateProgress(20, "enum sheet creating...")

                # enum worksheet부터 생성하자...
                # enum sheet에 있는 Enum CellRange
                enum_cell_range = self.__generate_enum_worksheet(workbook, self.__enum_data, self.__schema_data)

                if worker: worker.updateProgress(60, "data sheet sync...")

                # 갱신 된 enum값이 생길수 있으므로 sync를 맞춰 준다.
                self.__format_sync_datasheet(workbook, self.__enum_data, self.__schema_data, enum_cell_range)

                if worker: worker.updateProgress(80, "data sheet value validate checking...")

                # validation도 없데이트가 되었으므로 기존 값이 정상인지 체크
                self.__check_data_for_validation(workbook, self.__enum_data, self.__schema_data)

                if worker: workbook.save(target_excel_data_path)

                if worker: worker.updateProgress(100, "{0} format sync complete".format(filename))
            except:
                raise
            finally:
                if workbook is not None:
                    workbook.close()

        return work

    def new_excel_data_async(self, new_data_path:str):
        
        if new_data_path.endswith('.xlsx') == False:
            raise Exception('{0} is not .xlsx file'.format(new_data_path))

        new_data_path = convert_path(new_data_path)
        mkdir_path(new_data_path)

        worker = ProgressWorker(self.__new_excel_data(new_data_path))
        worker.start()

        return worker

    def new_excel_data_sync(self, new_data_path:str):

        if new_data_path.endswith('.xlsx') == False:
            raise Exception('{0} is not .xlsx file'.format(new_data_path))

        new_data_path = convert_path(new_data_path)
        mkdir_path(new_data_path)

        if os.path.exists(new_data_path):
            raise Exception(f'{new_data_path} is exist')

        return self.__new_excel_data(new_data_path)(None)

    def __new_excel_data(self, new_excel_data_path:str):

        def work(worker:ProgressWorker):
            
            filename = os.path.basename(new_excel_data_path)

            try:
                if worker: worker.updateProgress(0, "{0} creating...".format(filename))

                workbook = Workbook()
                workbook.remove(workbook.active)

                if worker: worker.updateProgress(20, "enum sheet creating...")

                # enum worksheet부터 생성하자...
                # enum sheet에 있는 Enum CellRange
                enum_cell_range = self.__generate_enum_worksheet(workbook, self.__enum_data, self.__schema_data)

                if worker: worker.updateProgress(60, "data sheet creating...")
                
                #DATA 시트가 없으면 새로 만든다.
                if 'DATA' not in workbook.sheetnames:
                    self.__new_data_worksheet(workbook, self.__schema_data)

                self.__format_sync_datasheet(workbook, self.__enum_data, self.__schema_data, enum_cell_range)

                workbook.move_sheet("ENUM", 1)

                workbook.save(new_excel_data_path)
                workbook.close()

                if worker: worker.updateProgress(100, "{0} create complete".format(filename))
            except Exception as e:
                raise e
            finally:
                if workbook is not None:
                    workbook.close()

        return work

    def __new_data_worksheet(self, workbook:Workbook, schema_data:ExcelSchemaData):
        if 'DATA' in workbook.sheetnames:
            workbook.remove(workbook['DATA'])

        data_worksheet = workbook.create_sheet('DATA')

        column_field_group = [
            'name',
            'type',
            'default',
            'title',
            'comment'
        ]

        for idx, column_field in enumerate(column_field_group):
            cell = data_worksheet.cell(2+idx, 1)
            cell.value = column_field

        start_row = 2
        for idx, column_field in enumerate(schema_data.get_fields()):
            cell = data_worksheet.cell(start_row, 2 + idx)
            cell.value = column_field.name

            cell = data_worksheet.cell(start_row + 1, 2 + idx)
            cell.value = column_field.type

            cell = data_worksheet.cell(start_row + 2, 2 + idx)
            cell.value = column_field.get_native_default()

            cell = data_worksheet.cell(start_row + 3, 2 + idx)
            cell.value = column_field.title

            cell = data_worksheet.cell(start_row + 4, 2 + idx)
            cell.value = column_field.comment

        # empty Cell들 생성
        for row in data_worksheet.iter_rows(min_row=7, max_row=self.__default_row_count, min_col=1, max_col=schema_data.get_field_count()):
            for cell in row:
                cell.value = ''
                        
    def __format_sync_datasheet(self, workbook:Workbook, enum_data:EnumMetaData, schema_data:ExcelSchemaData, enum_cell_range:dict):
        data_sheet = workbook['DATA']
        if data_sheet is None:
            raise Exception('DATA sheet is not exists')

        data_sheet.data_validations = DataValidationList()

        sheet_max_row = data_sheet.max_row if data_sheet.max_row > self.__default_row_count else self.__default_row_count
        sheet_max_column = data_sheet.max_column

        # (2, 2) ~ (2, sheet_max_column) 범위 체크
        # schema_field 목록
        for column_idx in range(2, sheet_max_column + 1):

            # 정상적인 필드 이름 체크
            schema_field_name:str = data_sheet.cell(2, column_idx).value
            schema_field:ExcelSchemaField = self.__get_valid_field_name(schema_field_name, schema_data)
            if schema_field is None:
                continue

            # type이 Enum Type이면 Comment 처리 한다
            type_cell = data_sheet.cell(3, column_idx)
            type_cell.comment = None

            if schema_field.get_nativetype() == Enum:
                
                if enum_data.is_exist_enum(schema_field.type) is False:
                    raise Exception('{0} {1} enum type is not exist'.format(schema_field.name, schema_field.type))

                enum_meta_info:EnumMetaInfo = enum_data.get_enum_meta_info(schema_field.type)
                comment = enum_meta_info.to_excel_comment()
                
                type_cell.comment = Comment(comment, '', width=400, height=400)

            # DataValidation 생성
            data_validation = self.__get_validation(schema_field, enum_cell_range)
            if data_validation:

                data_sheet.add_data_validation(data_validation)

                # 실제 Data가 시작하는 행은 7번 행부터 시작한다.
                for row_idx in range(7, sheet_max_row):
                    data_cell = data_sheet.cell(row_idx, column_idx)
                    data_validation.add(data_cell)

    def __check_data_for_validation(self, workbook:Workbook, enum_data:EnumMetaData, schema_data:ExcelSchemaData):
        data_sheet = workbook['DATA']
        if data_sheet is None:
            raise Exception('DATA sheet is not exists')

        for column_idx in range(2, data_sheet.max_column + 1):

            schema_cell:Cell = data_sheet.cell(2, column_idx)
            schema_cell_name:str = schema_cell.value if schema_cell.value is not None else ''

            schema_field:ExcelSchemaField = schema_data.find_schema_field(schema_cell_name)
            if schema_field is None:
                continue

            check_func = None
            enum_meta_data:EnumMetaInfo = None

            # schema_type에 맞는 value range
            if schema_field.get_nativetype() == numpy.int32:
                check_func = ef_util.is_valid_value_int32
            elif schema_field.get_nativetype() == numpy.uint32:
                check_func = ef_util.is_valid_value_uint32
            elif schema_field.get_nativetype() == numpy.float32:
                check_func = ef_util.is_valid_value_float32
            elif schema_field.get_nativetype() == bool:
                check_func = ef_util.is_valid_value_bool
            elif schema_field.get_nativetype() == str:
                check_func = lambda v, *args : True
            elif schema_field.get_nativetype() == Enum:
                check_func = ef_util.is_valid_value_enum

                enum_meta_data = enum_data.get_enum_meta_info(schema_field.type)

            for row_idx in range(7, data_sheet.max_row + 1):
                
                value_cell = data_sheet.cell(row_idx, column_idx)

                if check_func is not None:
                    if check_func(value_cell.value, enum_meta_data) == False:
                        value_cell.value = ''

    def __generate_enum_worksheet(self, workbook:Workbook, enum_data:EnumMetaData, schema_data:ExcelSchemaData):
        if 'ENUM' in workbook.sheetnames:
            workbook.remove(workbook['ENUM'])

        enum_worksheet = workbook.create_sheet('ENUM', )
        enum_worksheet.sheet_state = 'hidden'

        # Schema에서 사용되는 Enum 타입들 조사
        use_enum = {}
        for schema_field in schema_data.get_fields():
            if schema_field.get_nativetype() != Enum:
                continue
        
            if enum_data.is_exist_enum(schema_field.type) == False:
                raise Exception('enum \'{0}\' type \'{1}\' not found'.format(schema_field.name, schema_field.type))

            use_enum[schema_field.type] = enum_data.get_enum_meta_info(schema_field.type)

        enum_cell_range = {}
        use_enum_key_list = use_enum.keys()
        for idx, enum_key in enumerate(use_enum_key_list):
            
            # 1열에 Enum Name
            cell = enum_worksheet.cell(1, idx + 1)
            cell.value = enum_key

            # 2열 부터 Enum의 Value들
            enum_meta:EnumMetaInfo = use_enum[enum_key]
            for field_idx, enum_field in enumerate(enum_meta.enum_fields):
                cell = enum_worksheet.cell(2+field_idx, idx + 1)
                cell.value = enum_field.field_name

            startIndex:str = '${0}${1}'.format(chr(ord('A') + idx), 2)
            endIndex:str = '${0}${1}'.format(chr(ord('A') + idx), 2 + len(enum_meta.enum_fields) - 1)
            
            enum_cell_range[enum_key] = (startIndex, endIndex)
        
        return enum_cell_range

    def __get_valid_field_name(self, schema_field_name:str, schema_data:ExcelSchemaData):
        schema_field_name = schema_field_name.strip()
        if schema_field_name.startswith('//') == True:
            return None

        return schema_data.find_schema_field(schema_field_name)

    def __get_validation(self, schema_field:ExcelSchemaField, enum_cell_range:dict):
        
        if schema_field.get_nativetype() == numpy.int32:
            int32iinfo = numpy.iinfo(numpy.int32)
            return DataValidation(type='decimal', operator='between', formula1=int32iinfo.min, formula2=int32iinfo.max)
        
        if schema_field.get_nativetype() == numpy.uint32:
            uint32iinfo = numpy.iinfo(numpy.uint32)
            return DataValidation(type='decimal', operator='between', formula1=uint32iinfo.min, formula2=uint32iinfo.max)

        if schema_field.get_nativetype() == numpy.float32:
            float32iinfo = numpy.finfo(numpy.float32)
            return DataValidation(type='decimal', operator='between', formula1=float32iinfo.min, formula2=float32iinfo.max)

        if schema_field.get_nativetype() == str:
            return DataValidation(type='textLength', operator='lessThanOrEqual', formula1=1024)

        if schema_field.get_nativetype() == bool:
            return DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True)

        if schema_field.get_nativetype() == Enum:

            if schema_field.type not in enum_cell_range:
                raise Exception('enum_cell_rage not exit {0}'.format(schema_field.type))

            enum_range = enum_cell_range[schema_field.type]
            formular = '=ENUM!{0}:{1}'.format(enum_range[0], enum_range[1])

            data_validation = DataValidation(type='list', operator='between', formula1=formular)
            return data_validation
        
        return None