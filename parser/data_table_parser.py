
import os
from enum import Enum
from numpy import False_

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook

from core.progress_woker import ProgressWorker

from parser.schema_table_parser import ExcelSchemaData, ExcelSchemaField
from parser.enum_define_parser import EnumMetaData
from parser.define import is_valid_primitive_value, is_valid_enum_value

class ExcelData:

    __excel_schema_data:ExcelSchemaData = None

    __data_rows = [] # list<list<object>>
    __data_column_fields = [] # list<str>

    def __init__(self, excel_schema_data:ExcelSchemaData, data_column_fileds, data_rows):
        self.__excel_schema_data = excel_schema_data
        self.__data_column_fields = data_column_fileds
        self.__data_rows = data_rows

    def get_excel_schema_data(self) -> ExcelSchemaData:
        return self.__excel_schema_data

    def get_row_count(self):
        return len(self.__data_rows)

    def get_row(self, index):
        if index >= self.get_row_count():
            return None

        return self.__data_rows[index]

    def get_rows(self):
        for row in self.__data_rows:
            yield row

    def get_column_mapping_rows(self):
        
        for row in self.__data_rows:

            mapping_row = []
            for schema_field in self.__excel_schema_data.get_fields():
                index = self.get_column_field_index(schema_field.name)
                if index == -1:
                    continue

                mapping_row.append((schema_field.name, row[index]))

            yield mapping_row

    def get_column_fields(self):
        return self.__data_column_fields

    def get_column_field_index(self, field_name:str):
        return self.__data_column_fields.index(field_name)

    def merge(self, excel_data):

        if not isinstance(excel_data, ExcelData):
            return ValueError('excel_data is not ExcelData')

        other_schema_data:ExcelSchemaData = excel_data.get_excel_schema_data()
        if self.get_excel_schema_data() != other_schema_data:
            raise ValueError(f'another data[schema : {other_schema_data.schema_name}] is different')

        if len(self.get_column_fields()) != len(excel_data.get_column_fields()):
            raise ValueError('column fields not different')

        for idx, field_name in enumerate(self.get_column_fields()):
            another_field_name = excel_data.get_column_fields()[idx]

            if field_name != another_field_name:
                raise ValueError('column fields not different')

        for row in excel_data.get_rows():
            self.__data_rows.append(row)

class ExcelDataParser:

    __is_convert_sql_value:bool = False
    __filepath:str = ''

    def __init__(self, filepath:str, *, is_convert_sql_value:bool):
        self.__filepath = filepath
        self.__is_convert_sql_value = is_convert_sql_value

        if os.path.exists(filepath) == False:
            raise FileNotFoundError('{0} is not found'.format(filepath))

        if self.__is_excel_file(filepath) == False:
            raise Exception('{0} is not excel file'.format(filepath))

    def parsing_async(self, excel_schmea_data:ExcelSchemaData, enum_data:EnumMetaData):

        worker = ProgressWorker(self.__parsing(excel_schmea_data, enum_data, self.__filepath))
        worker.start()

        return worker

    def parsing_sync(self, excel_schmea_data:ExcelSchemaData, enum_data:EnumMetaData):

        return self.__parsing(excel_schmea_data, enum_data, self.__filepath)(None)

    def __parsing(self, excel_schmea_data:ExcelSchemaData, enum_data:EnumMetaData, filepath:str):
        
        def work(worker:ProgressWorker):

            filename = os.path.basename(filepath)

            try:

                if worker: worker.updateProgress(0, "{0} loading...".format(filename))

                workbook:Workbook = load_workbook(filename = filepath)
                if 'DATA' not in workbook.sheetnames:
                    raise Exception('not found \'DATA\' sheet')

                worksheet:Worksheet = workbook['DATA']

                if worker: worker.updateProgress(20, "column parsing...")

                # data에 있는 schema Column 조사 하기
                column_fileds = self.__parse_data_columns(worksheet)

                # None 정규화 하기
                data_column_fiels = list(filter(lambda item : item is not None , column_fileds))
                data_rows = self.__parse_data_rows(worker, worksheet, column_fileds, excel_schmea_data, enum_data)

                # ExcelData 구성
                excel_data = ExcelData(excel_schmea_data, data_column_fiels, data_rows)

                if worker: worker.updateProgress(100, "{0} load complete...".format(filename))

                return excel_data

            except:
                raise
            finally:
                if workbook is not None:
                    workbook.close()

        return work

    def __parse_data_columns(self, workwsheet:Worksheet):
        max_column = workwsheet.max_column

        data_columns = []

        for colum_index in range(2, max_column + 1):

            cell = workwsheet.cell(2, colum_index)
            if cell is None or len(cell.value) == 0:
                data_columns.append(None)
                continue

            column_field_name = cell.value
            column_field_name = column_field_name.strip()

            if column_field_name.startswith('//') is True:
                data_columns.append(None)
                continue

            data_columns.append(column_field_name)

        return data_columns

    def __parse_data_rows(self, worker:ProgressWorker, workwsheet:Worksheet, column_order_list:list, schema_data:ExcelSchemaData, enum_data:EnumMetaData):
        
        start_row = 7
        max_row = workwsheet.max_row
        max_column = workwsheet.max_column

        # SchemaFiled를 조사 하기 위해.. {index, SchemaField}
        # 0 인덱스부터 시작해야 한다.
        column_index_mapping = {}
        for column_index in range(2, max_column + 1):

            if column_index not in column_index_mapping:

                index = column_index - 2
                column_field_name = column_order_list[index]
                column_index_mapping[index] = schema_data.find_schema_field(column_field_name)

        data_rows = []
        
        system_max_row = (max_row + 1 - start_row)
        progress_row_unit = system_max_row // 10
        progress_row_unit = progress_row_unit if progress_row_unit > 0 else 1
        
        for row_index in range(start_row, max_row + 1):

            if worker:
                system_curr_row = row_index - start_row
                curr_progress = 30 + (float(system_curr_row) / float(system_max_row)) * 70
                curr_progress = int(curr_progress)

                if system_curr_row % progress_row_unit == 0:
                    worker.updateProgress(curr_progress, "row {0} parsing...".format(system_curr_row))

            # 무시 체크
            cell = workwsheet.cell(row_index, 1)
            cell_str = str(cell.value if cell is not None else '')
            if cell_str.strip().startswith('//') is True:
                continue

            is_empty_cell = False
            # priamryKey 해당 하는 cell이 정상인가??
            for column_index in range(2, max_column + 1):
                
                # schemaFiled를 가져온다.
                column_schema_field:ExcelSchemaField = column_index_mapping[column_index - 2]
                if column_schema_field is None:
                    continue

                if column_schema_field.primary is False:
                    continue

                cell = workwsheet.cell(row_index, column_index)
                value = cell.value if cell is not None else None

                is_empty_cell = True if value is None else False

                if is_empty_cell is True:
                    break

            if is_empty_cell is True:
                continue

            # row 데이터 추출
            row_bundle = []
            for column_index in range(2, max_column + 1):

                # schemaFiled를 가져온다.
                column_schema_field:ExcelSchemaField = column_index_mapping[column_index - 2]
                if column_schema_field is None:
                    continue

                cell = workwsheet.cell(row_index, column_index)
                value = cell.value if cell is not None else None
                if value is None:
                    value = str(column_schema_field.get_native_default())
                elif isinstance(value, str) and len(value) == 0:
                    value = str(column_schema_field.get_native_default())
                else:
                    value = str(value)

                row_bundle.append(value)

            # row 데이터 값이 정상인가???
            for idx, value in enumerate(row_bundle):
                column_schema_field = column_index_mapping[idx]
                value:str = value

                if column_schema_field is None:
                    continue

                is_valid = is_valid_primitive_value(column_schema_field, value)
                if not is_valid:
                    is_valid = is_valid_enum_value(column_schema_field, enum_data, value)

                if not is_valid:
                    raise ValueError(f'error valid: [row:{row_index}] [column:{column_schema_field.name}] [type:{column_schema_field.type}] [value:{value}]')

            # row 데이터 sql value로 치환
            if self.__is_convert_sql_value:
                for idx, value in enumerate(row_bundle):
                    column_schema_field = column_index_mapping[idx]
                    value:str = value

                    row_bundle[idx] = column_schema_field.convert_sql_value(value, enum_data)


            data_rows.append(row_bundle)

        return data_rows

    def __is_excel_file(self, filepath):

        filename:str = os.path.basename(filepath)
        return True if filename.endswith('.xlsx') else False