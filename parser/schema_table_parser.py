import os
import numpy

from enum import Enum
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.progress_woker import ProgressWorker
from parser.enum_define_parser import EnumMetaData
from parser.define import SUPPORT_PRIMITIVE_TYPE

class ExcelSchemaField:
    name:str = ''
    type:str = ''
    default:str = ''
    title:str = ''
    comment:str = ''
    primary:bool = False

    def get_nativetype(self):

        if len(self.type) == 0:
            return None
        
        lower_type = self.type.lower()

        if lower_type == 'uint':
            return numpy.uint32

        if lower_type == 'int':
            return numpy.int32

        if lower_type == 'bool':
            return bool
        
        if lower_type == 'float':
            return numpy.float32

        if lower_type == 'string':
            return str

        return Enum

    def get_sqlitetype(self):

        native_type = self.get_nativetype()
        if native_type is numpy.int32 or\
            native_type is numpy.uint32 or\
            native_type is bool or\
            native_type is Enum:
            return 'Integer'

        if native_type is numpy.float32:
            return 'Real'

        if native_type is str:
            return 'Text'
    
        return None

    def get_native_default(self):

        native_type = self.get_nativetype()
        if native_type is numpy.uint32:
            return numpy.uint32(self.default)
        
        if native_type is numpy.int32:
            return numpy.int32(self.default)

        if native_type is numpy.float32:
            return numpy.float32(self.default)

        if native_type is bool:
            return True if self.default.lower() == 'true' else False

        return self.default

    def convert_sql_value(self, value:str, enum_data:EnumMetaData):

        if self.get_nativetype() is numpy.int32:
            return int(numpy.int32(value))
    
        if self.get_nativetype() is numpy.uint32:
            return int(numpy.uint32(value))

        if self.get_nativetype() is numpy.float32:
            return float(numpy.float32(value))

        if self.get_nativetype() is numpy.uint32:
            return int(numpy.uint32(value))

        if self.get_nativetype() is bool:
            return 1 if value.lower() == 'true' else 0
    
        if self.get_nativetype() is Enum:
            enum_meta_info = enum_data.get_enum_meta_info(self.type)
            enum_field = enum_meta_info.get_enum_field(value)
            return enum_field.field_value
        
        return value

    def __str__(self):

        output = '[name : {0}]'.format(self.name);
        output += '[type : {0}]'.format(self.type);
        output += '[default : {0}]'.format(self.default);
        output += '[title : {0}]'.format(self.title);
        output += '[comment : {0}]'.format(self.comment);
        output += '[primary : {0}]'.format(self.primary);

        return output;

    def __eq__(self, o: object):
        
        if not isinstance(o, ExcelSchemaField):
            return False
        
        if self.name != o.name:
            return False

        if self.type != o.type:
            return False

        if self.primary != o.primary:
            return False

        return True

    def __ne__(self, o: object):
        return not self.__eq__(o)

class ExcelSchemaData:
    schema_name:str = ''
    table_name:str = ''
    db_name:str = ''

    # list[SchemaFiled]
    __fields:list = []

    def __init__(self, schema_name:str, fields):
        self.__fields = fields
        self.schema_name = schema_name
        self.table_name = 'Tbl' + schema_name.capitalize()
        self.db_name = schema_name.capitalize()

    def find_schema_field(self, fieldname:str):

        if len(self.__fields) == 0:
            return None

        def _filter_func(item:ExcelSchemaField):
            if item.name.lower() == fieldname.lower():
                return item
            return None
            
        result = filter(_filter_func, self.__fields)
        result = list(result)
        if len(result) == 0:
            return None

        return result[0]

    def isprimary(self, fieldname:str):

        if len(self.__fields) == 0:
            return False

        schemafield:ExcelSchemaField = self.find_schema_field(fieldname)
        if schemafield is None:
            return False
        
        return schemafield.primary

    def get_field_count(self) -> int:
        return len(self.__fields)

    def get_fields(self) -> ExcelSchemaField:
        for field in self.__fields:
            yield field

    def get_filed(self, index:int):
        if index >= len(self.__fields):
            return None

        return self.__fields[index]

    def __str__(self):
        print('schema_name = {0}'.format(self.schema_name))
        print('table_name = {0}'.format(self.table_name))
        print('db_name = {0}'.format(self.db_name))
        for field in self.__fields:
            print('[{0}] - {1}'.format(self.__fields.index(field), field))

    def __eq__(self, o: object):
        if not isinstance(o, ExcelSchemaData):
            return False
        
        if self.schema_name != o.schema_name:
            return False
        
        if self.table_name != o.table_name:
            return False
        
        if self.db_name != o.db_name:
            return False

        if self.get_field_count() != o.get_field_count():
            return False

        for index in range(0, self.get_field_count()):
            if self.get_filed(index) != o.get_filed(index):
                return False

        return True

    def __ne__(self, o: object):
        return not self.__eq__(o)

class ExcelSchemaParser:

    __filepath:str = ''

    def __init__(self, filepath:str):
        self.__filepath = filepath

        if os.path.exists(filepath) == False:
            raise FileNotFoundError('{0} not found'.format(filepath))

        if self.__is_schema_file(filepath) == False:
            raise Exception('filepath is not schema file')

    def parsing_async(self, enum_data:EnumMetaData):

        worker = ProgressWorker(self.__parsing(self.__filepath, enum_data))
        worker.start()

        return worker

    def parsing_sync(self, enum_data:EnumMetaData):
        return self.__parsing(self.__filepath, enum_data)(None)

    def __parsing(self, filepath:str, enum_data:EnumMetaData):

        def work(worker:ProgressWorker):

            filename = os.path.basename(filepath)

            if worker: worker.updateProgress(0, "{0} loading..".format(filename))

            try:

                workbook:Workbook = load_workbook(filename = filepath)
                if 'SCHEMA' not in workbook.sheetnames:
                    raise Exception('not found \'SCHEMA\' sheet')

                if worker: worker.updateProgress(50, "{0} load complete".format(filename))
            
                worksheet:Worksheet = workbook['SCHEMA']
            
                schema_name:str = os.path.basename(filepath)
                findindex = schema_name.lower().index('.schema.xlsx')
                schema_name = schema_name[0: findindex]

                # schema 첫 행 정보 구성하기
                # <string, int> = <schemafieldname, filed가 위치하는 Column Index>
                column_oridinal = self.__set_column_ordinal(worksheet)

                # schema 첫행 스키마 구성 요소들이 정상인가?? (name, type, default, title, comment, primary)
                if self.__is_all_exist_schema_fields(column_oridinal) == False:
                    raise Exception('not all exist schema fileds')

                # feild 정보 파싱
                schema_fields = self.__get_schema_fields(worksheet, column_oridinal)

                # field type 유효성 체크
                self.__check_valid_schema_field(schema_fields, enum_data)

                # SchemaData 구성
                schema_data = ExcelSchemaData(schema_name, schema_fields)

                # column_filed 유효성 체크
                for field in schema_data.get_fields():
                    f:ExcelSchemaField = field
                    f.type

                if worker: worker.updateProgress(100, "{0} schema data make complete".format(filename))

                return schema_data
            except:
                raise;
            finally:
                if workbook is not None:
                    workbook.close()

        return work


    def __is_schema_file(self, filepath):

        filename:str = os.path.basename(filepath)
        filename = filename.lower()
        if filename.endswith('.xlsx') == False:
            return False

        if filename.endswith('.schema.xlsx') == False:
            return False

        return True


    # field의 특성 가져온다. 첫줄에 정의 되어있음..
    # Schema 정보의 field들을 수집
    def __set_column_ordinal(self, worksheet:Worksheet):
        column_oridinal = {}

        max_column = worksheet.max_column
        for column_index in range(1, max_column+1):
            cell = worksheet.cell(1, column_index)
            if cell is None:
                continue

            column_oridinal[cell.value.lower()] = column_index

        return column_oridinal

    def __get_schema_fields(self, worksheet:Worksheet, column_oridinal:dict):

        schema_fields = []

        max_row = worksheet.max_row
        for row_index in range(2, max_row + 1):
            
            schema_field = ExcelSchemaField()

            # default 필드를 제외하고 모두 조사
            column_keys = [key for key in column_oridinal.keys() if key != 'default']
            for column_key in column_keys:

                column_index = column_oridinal[column_key]

                cell = worksheet.cell(row_index, column_index)

                # Name, Type에 해당하는 cell은 무조건 값이 있어야 한다.
                if column_key == 'name' or column_key == 'type':
                    if cell is None or len(cell.value) == 0:
                        raise Exception('{0} field is empty'.format(column_key))

                # cell.value에 값이 없으면 기본값을 채워주자...
                default_value = str(cell.value) if cell.value is not None else None
                if default_value is None:
                    
                    attr = getattr(schema_field, column_key)
                    if type(attr) is str:
                        default_value = ''
                    elif type(attr) is bool:
                        default_value = 'False'

                # primary 값 native로 변경
                if column_key == 'primary':
                    default_value = True if default_value.lower() == 'true' else False
            
                setattr(schema_field, column_key, default_value)

            # default 조사
            column_keys = [key for key in column_oridinal.keys() if key == 'default']
            for column_key in column_keys:
                
                column_index = column_oridinal[column_key]
                cell = worksheet.cell(row_index, column_index)

                default_value = str(cell.value) if cell.value is not None else ''
                if len(default_value) != 0:
                    setattr(schema_field, column_key, default_value)
                    continue

                if schema_field.get_nativetype() == numpy.int32:
                    default_value = '0'
                elif schema_field.get_nativetype() == numpy.uint32:
                    default_value = '0'
                elif schema_field.get_nativetype() == numpy.float32:
                    default_value = '0'
                elif schema_field.get_nativetype() == bool:
                    default_value = 'False'
                elif schema_field.get_nativetype() == Enum:
                    default_value = 'None'
                else:
                    default_value = ''

                setattr(schema_field, column_key, default_value)

            schema_fields.append(schema_field)

        return schema_fields

    # schema field 들이 정상적으로 구성 되어 있나??
    def __is_all_exist_schema_fields(self, column_oridinal:dict):

        check_schema_field = ExcelSchemaField()

        for name in column_oridinal.keys():
            if hasattr(check_schema_field, name) == False:
                return False

        return True

    # schema field 구성 요소가 정상적인가??
    def __check_valid_schema_field(self, schema_fields:list, enum_data:EnumMetaData):
        
        for schema_field in schema_fields:
            schema_field:ExcelSchemaField = schema_field

            if not schema_field.type.lower() in SUPPORT_PRIMITIVE_TYPE:

                # enum 타입에 있는지 확인
                if not enum_data.is_exist_enum(schema_field.type):
                    raise ValueError(f'{schema_field.name} type[{schema_field.type}] not support')